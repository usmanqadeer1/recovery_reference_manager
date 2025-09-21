from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from .models import Reference, MediaFile, Category
from .forms import ReferenceForm, MediaFileForm, ExcelImportForm

def home(request):
    """Home page showing all references"""
    references = Reference.objects.all().order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        references = references.filter(
            Q(reference_number__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(references, 10)  # Show 10 references per page
    page_number = request.GET.get('page')
    references = paginator.get_page(page_number)
    
    context = {
        'references': references,
        'search_query': search_query,
    }
    return render(request, 'references/home.html', context)

def reference_detail(request, reference_id):
    """Detail view for a specific reference"""
    reference = get_object_or_404(Reference, id=reference_id)
    media_files = reference.media_files.all().order_by('-uploaded_at')
    
    context = {
        'reference': reference,
        'media_files': media_files,
    }
    return render(request, 'references/reference_detail.html', context)

def add_reference(request):
    """Add a new reference"""
    if request.method == 'POST':
        form = ReferenceForm(request.POST)
        if form.is_valid():
            reference = form.save()
            messages.success(request, f'Reference "{reference.reference_number}" added successfully!')
            return redirect('reference_detail', reference_id=reference.id)
    else:
        form = ReferenceForm()
    
    context = {'form': form}
    return render(request, 'references/add_reference.html', context)

def add_media(request, reference_id):
    """Add media files to a reference"""
    reference = get_object_or_404(Reference, id=reference_id)
    
    if request.method == 'POST':
        form = MediaFileForm(request.POST, request.FILES)
        if form.is_valid():
            media_file = form.save(commit=False)
            media_file.reference = reference
            media_file.save()
            form.save_m2m()  # Save many-to-many relationships (categories)
            messages.success(request, 'Media file added successfully!')
            return redirect('reference_detail', reference_id=reference.id)
    else:
        form = MediaFileForm()
    
    context = {
        'form': form,
        'reference': reference,
    }
    return render(request, 'references/add_media.html', context)

def edit_reference(request, reference_id):
    """Edit an existing reference"""
    reference = get_object_or_404(Reference, id=reference_id)
    
    if request.method == 'POST':
        form = ReferenceForm(request.POST, instance=reference)
        if form.is_valid():
            form.save()
            messages.success(request, 'Reference updated successfully!')
            return redirect('reference_detail', reference_id=reference.id)
    else:
        form = ReferenceForm(instance=reference)
    
    context = {
        'form': form,
        'reference': reference,
    }
    return render(request, 'references/edit_reference.html', context)

def import_excel(request):
    """Import references from Excel file"""
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                from openpyxl import load_workbook
                
                # Read Excel file
                workbook = load_workbook(excel_file)
                worksheet = workbook.active
                
                # Get headers from first row
                headers = []
                for cell in worksheet[1]:
                    headers.append(cell.value.lower().strip() if cell.value else '')
                
                # Check required columns
                required_columns = ['reference_number', 'title', 'tariff', 'age', 'amount', 'description']
                missing_columns = [col for col in required_columns if col not in headers]
                
                if missing_columns:
                    messages.error(request, f'Missing required columns: {", ".join(missing_columns)}')
                    return render(request, 'references/import_excel.html', {'form': form})
                
                # Get column indices
                col_indices = {col: headers.index(col) for col in required_columns}
                
                # Process each row
                created_count = 0
                updated_count = 0
                errors = []
                
                for row_num in range(2, worksheet.max_row + 1):
                    try:
                        row = worksheet[row_num]
                        
                        # Extract data from each column
                        reference_number = str(row[col_indices['reference_number']].value or '').strip()
                        title = str(row[col_indices['title']].value or '').strip()
                        tariff = str(row[col_indices['tariff']].value or '').strip() or None
                        
                        # Handle age
                        age_value = row[col_indices['age']].value
                        age = int(age_value) if age_value is not None and str(age_value).strip() != '' else None
                        
                        # Handle amount
                        amount_value = row[col_indices['amount']].value
                        amount = float(amount_value) if amount_value is not None and str(amount_value).strip() != '' else None
                        
                        description = str(row[col_indices['description']].value or '').strip() or None
                        
                        if not reference_number or not title:
                            errors.append(f'Row {row_num}: Reference number and title are required')
                            continue
                        
                        # Check if reference exists
                        reference, created = Reference.objects.get_or_create(
                            reference_number=reference_number,
                            defaults={
                                'title': title,
                                'tariff': tariff,
                                'age': age,
                                'amount': amount,
                                'description': description
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            # Update existing reference
                            reference.title = title
                            reference.tariff = tariff
                            reference.age = age
                            reference.amount = amount
                            reference.description = description
                            reference.save()
                            updated_count += 1
                            
                    except Exception as e:
                        errors.append(f'Row {row_num}: {str(e)}')
                        continue
                
                # Show results
                if created_count > 0 or updated_count > 0:
                    success_msg = f'Successfully processed {created_count + updated_count} references: {created_count} created, {updated_count} updated.'
                    messages.success(request, success_msg)
                
                if errors:
                    error_msg = f'Errors occurred in {len(errors)} rows: ' + '; '.join(errors[:5])
                    if len(errors) > 5:
                        error_msg += f' ... and {len(errors) - 5} more errors.'
                    messages.warning(request, error_msg)
                
                return redirect('home')
                
            except Exception as e:
                messages.error(request, f'Error reading Excel file: {str(e)}')
    else:
        form = ExcelImportForm()
    
    context = {'form': form}
    return render(request, 'references/import_excel.html', context)

def download_template(request):
    """Download Excel template for import"""
    try:
        from openpyxl import Workbook
        import io
        
        # Create a new workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "References"
        
        # Add headers
        headers = ['reference_number', 'title', 'tariff', 'age', 'amount', 'description']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Add sample data
        sample_data = [
            ['REF001', 'Sample Reference 1', 'T1', 25, 100000.50, 'Sample description 1'],
            ['REF002', 'Sample Reference 2', 'T2', 30, 200000.75, 'Sample description 2'],
            ['REF003', 'Sample Reference 3', 'T3', 35, 300000.00, 'Sample description 3']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                worksheet.cell(row=row, column=col, value=value)
        
        # Create Excel file in memory
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reference_template.xlsx"'
        
        return response
    except ImportError:
        messages.error(request, 'Excel functionality not available. Please install required packages.')
        return redirect('home')
